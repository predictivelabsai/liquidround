"""Valuation Simulator — 4-method company valuation with WACC, industry
multiples (Damodaran), equity bridge, Plotly chart, and XLS export.

Ported from pehero/chat/valuation.py for LiquidRound's navy+amber palette.
Uses COMPANY_DB_SCHEMA env var (default 'pehero') for company data.

/app/valuation              -> simulator page
/app/valuation/search       -> autocomplete JSON
/app/valuation/export       -> downloadable XLS workbook
"""

from __future__ import annotations

import csv
import io
import json
import os
from pathlib import Path

from fasthtml.common import (
    Html, Head, Body, Meta, Title, Link, Script, NotStr,
    Div, Span, H2, H3, H4, P, A, Button, Form, Input, Select, Option, Label,
)
from fasthtml.core import APIRouter
from starlette.requests import Request
from starlette.responses import JSONResponse, Response

from utils.database import get_conn

ar = APIRouter()

DB_SCHEMA = os.getenv("COMPANY_DB_SCHEMA", "pehero")
DATA = Path(__file__).resolve().parent / "data"


def _load_revenue_multiples() -> dict[str, float]:
    """Load from DB table first, fall back to CSV."""
    out = {}
    try:
        rows = _fetch_all(
            f"SELECT industry_sector, sales_multiplier "
            f"FROM {DB_SCHEMA}.valuation_multipliers_revenue"
        )
        for r in rows:
            out[r["industry_sector"]] = float(r["sales_multiplier"])
        if out:
            return out
    except Exception:
        pass
    p = DATA / "valuation-multipliers - revenue.csv"
    if not p.exists():
        return out
    with open(p) as f:
        for row in csv.DictReader(f):
            try:
                out[row["industry_sector"]] = float(row["sales_multiplier"])
            except (ValueError, KeyError):
                pass
    return out


def _load_ebitda_multiples() -> dict[str, float]:
    """Load from DB table first, fall back to CSV."""
    out = {}
    try:
        rows = _fetch_all(
            f"SELECT industry_name, ev_ebitda, ev_ebit "
            f"FROM {DB_SCHEMA}.valuation_multipliers_ebitda"
        )
        for r in rows:
            name = r["industry_name"]
            ebitda = float(r["ev_ebitda"]) if r["ev_ebitda"] is not None else None
            ebit = float(r["ev_ebit"]) if r["ev_ebit"] is not None else None
            if ebitda is not None:
                out[name] = {"ebitda": ebitda, "ebit": ebit}
        if out:
            return out
    except Exception:
        pass
    p = DATA / "valuation-multipliers - ebidta.csv"
    if not p.exists():
        return out
    lines = p.read_text().splitlines()
    header_idx = None
    for i, line in enumerate(lines):
        if line.startswith("Industry Name,"):
            header_idx = i
            break
    if header_idx is None:
        return out
    reader = csv.DictReader(lines[header_idx:])
    for row in reader:
        name = row.get("Industry Name", "").strip()
        ev_ebitda = row.get("EV/EBITDA", "").strip()
        ev_ebit = row.get("EV/EBIT", "").strip()
        if not name:
            continue
        try:
            out[name] = {"ebitda": float(ev_ebitda), "ebit": float(ev_ebit) if ev_ebit and ev_ebit != "NA" else None}
        except ValueError:
            pass
    return out


REVENUE_MULT = _load_revenue_multiples()
EBITDA_MULT = _load_ebitda_multiples()
INDUSTRY_LIST = sorted(set(list(REVENUE_MULT.keys()) + list(EBITDA_MULT.keys())))

SECTOR_TO_INDUSTRY = {
    "healthcare":         "Hospitals/Healthcare Facilities",
    "software":           "Software (System & Application)",
    "industrials":        "Engineering/Construction",
    "financial_services": "Financial Svcs. (Non-bank & Insurance)",
    "business_services":  "Business & Consumer Services",
    "consumer":           "Retail (General)",
}


def _f(v) -> float:
    if v is None:
        return 0.0
    return float(v)


def _fetch_one(sql: str, params: tuple = ()):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            cols = [d[0] for d in cur.description]
            row = cur.fetchone()
            return dict(zip(cols, row)) if row else None


def _fetch_all(sql: str, params: tuple = ()):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            cols = [d[0] for d in cur.description]
            return [dict(zip(cols, r)) for r in cur.fetchall()]


def _metric_card(label: str, value: str):
    return Div(
        Span(label, cls="val-metric-label"),
        Span(value, cls="val-metric-value"),
        cls="val-metric-card",
    )


def _slider(id: str, label: str, default, min_val, max_val, step):
    return Div(
        Div(
            Label(label, cls="val-slider-label"),
            Span(str(default), id=f"{id}_val", cls="val-slider-val"),
            cls="val-slider-header",
        ),
        Input(type="range", id=id, name=id, cls="val-slider",
              min=str(min_val), max=str(max_val), step=str(step),
              value=str(default)),
        cls="val-slider-group",
    )


def _autocomplete_js() -> str:
    return """
(function() {
    var debounceTimer;
    var dropdown = document.getElementById("val-ac-results");
    var input = document.getElementById("val-search");
    if (!input || !dropdown) return;

    window.valAutocomplete = function(q) {
        clearTimeout(debounceTimer);
        if (q.length < 1) { dropdown.innerHTML = ""; dropdown.classList.remove("open"); return; }
        debounceTimer = setTimeout(function() {
            fetch("/app/valuation/search?q=" + encodeURIComponent(q))
                .then(function(r) { return r.json(); })
                .then(function(items) {
                    if (!items.length) {
                        dropdown.innerHTML = '<div class="val-ac-empty">No matches</div>';
                        dropdown.classList.add("open");
                        return;
                    }
                    dropdown.innerHTML = items.map(function(it) {
                        return '<div class="val-ac-item" onclick="valSelectCompany(\\'' + it.slug + '\\',\\'' + it.label.replace(/'/g, "\\\\'") + '\\')">' +
                            '<span class="val-ac-name">' + it.label + '</span>' +
                            '<span class="val-ac-sub">' + it.sub + '</span></div>';
                    }).join("");
                    dropdown.classList.add("open");
                });
        }, 180);
    };

    window.valSelectCompany = function(slug, label) {
        dropdown.innerHTML = "";
        dropdown.classList.remove("open");
        input.value = label;
        window.location.href = "/app/valuation?company=" + slug;
    };

    document.addEventListener("click", function(e) {
        if (!e.target.closest(".val-ac-wrap")) {
            dropdown.innerHTML = "";
            dropdown.classList.remove("open");
        }
    });
})();
"""


@ar.route("/app/valuation")
def valuation_home(sess, company: str = ""):
    from components.chat_shell import left_pane

    sym = "€"
    user_email = sess.get("email") if sess else None
    current_role = sess.get("role", "buyer") if sess else "buyer"

    selected = None
    if company:
        selected = _fetch_one(
            f"SELECT slug, name, sector, sub_sector, revenue_ltm, ebitda_ltm, "
            f"ebitda_margin, growth_rate, employees, hq_city, country, "
            f"enterprise_value "
            f"FROM {DB_SCHEMA}.companies WHERE slug = %s", (company,)
        )

    selected_label = ""
    if selected:
        selected_label = f"{selected['name']} — {sym}{_f(selected['revenue_ltm'])/1e6:.1f}M"

    _SAMPLE_SLUGS = [
        "gyvunu_ligonine", "eurovaistine", "eika_construction",
        "baltic_transline_uab", "vinted", "hegelmann_transporte",
    ]
    sample_companies = _fetch_all(
        f"SELECT slug, name FROM {DB_SCHEMA}.companies WHERE slug = ANY(%s)",
        (_SAMPLE_SLUGS,),
    )
    slug_order = {s: i for i, s in enumerate(_SAMPLE_SLUGS)}
    sample_companies.sort(key=lambda c: slug_order.get(c["slug"], 99))
    sample_chips = [
        A(c["name"], href=f"/app/valuation?company={c['slug']}", cls="val-chip")
        for c in sample_companies
    ]

    if not selected:
        selector = Div(
            Div(
                Span("◎", cls="val-hero-mark"),
                H2("Valuation Simulator", cls="val-hero-title"),
                P("Select a company to run a 4-method valuation", cls="val-hero-sub"),
                Div(
                    Div(
                        Input(type="text", id="val-search",
                              placeholder="Search by company name, city, or sector...",
                              cls="val-hero-input", value="", autocomplete="off",
                              oninput="valAutocomplete(this.value)",
                              onfocus="valAutocomplete(this.value)"),
                        Div(id="val-ac-results", cls="val-ac-dropdown"),
                        cls="val-ac-wrap",
                    ),
                    cls="val-hero-input-wrap",
                ),
                Div(*sample_chips, cls="val-chips") if sample_chips else None,
                cls="val-hero",
            ),
            cls="val-selector-wrap val-hero-wrap",
        )
    else:
        selector = Div(
            Div(
                Input(type="text", id="val-search",
                      placeholder="Search companies...",
                      cls="val-search-input", value=selected_label, autocomplete="off",
                      oninput="valAutocomplete(this.value)",
                      onfocus="valAutocomplete(this.value)"),
                Div(id="val-ac-results", cls="val-ac-dropdown"),
                cls="val-ac-wrap",
            ),
            cls="val-selector-wrap",
        )

    summary = Div(cls="val-summary", id="val-summary")
    simulator = Div(cls="val-simulator", id="val-simulator")
    sim_js = ""

    if selected:
        rev = _f(selected["revenue_ltm"])
        ebitda = _f(selected["ebitda_ltm"])
        margin = _f(selected["ebitda_margin"])
        growth = _f(selected["growth_rate"])
        employees = selected.get("employees") or 0
        sector = selected["sector"] or "healthcare"
        industry_default = SECTOR_TO_INDUSTRY.get(sector, "Business & Consumer Services")

        rev_mult_default = REVENUE_MULT.get(industry_default, 2.0)
        ebitda_entry = EBITDA_MULT.get(industry_default, {})
        ebitda_mult_default = ebitda_entry.get("ebitda", 10.0) if isinstance(ebitda_entry, dict) else 10.0
        ebit_mult_default = ebitda_entry.get("ebit", 12.0) if isinstance(ebitda_entry, dict) else 12.0
        if ebit_mult_default is None:
            ebit_mult_default = ebitda_mult_default * 1.3

        ebit = ebitda * 0.85 if ebitda > 0 else 0

        summary = Div(
            Div(
                H3(selected["name"], cls="val-company-name"),
                Span(f"{selected.get('hq_city') or ''} · {sector.replace('_', ' ').title()}"
                     + (f" · {employees} employees" if employees else ""),
                     cls="val-company-meta"),
                cls="val-company-header",
            ),
            Div(
                _metric_card("Revenue (LTM)", f"{sym}{rev/1e6:.2f}M"),
                _metric_card("EBITDA (LTM)", f"{sym}{ebitda/1e6:.2f}M"),
                _metric_card("EBIT", f"{sym}{ebit/1e6:.2f}M"),
                _metric_card("Margin", f"{margin:.1f}%"),
                _metric_card("Growth", f"{growth:+.1f}%"),
                _metric_card("Employees", f"{employees:,}" if employees else "—"),
                cls="val-metrics",
            ),
            cls="val-summary", id="val-summary",
        )

        industry_options = [Option(ind, value=ind, selected=ind == industry_default) for ind in INDUSTRY_LIST]

        multiples_row = Div(
            Div(
                H4("EV / Revenue Multiple", cls="val-method-title"),
                _slider("rev_multiple", "EV / Revenue", round(rev_mult_default, 1), 0.1, 25.0, 0.1),
                Div(Span("Enterprise Value", cls="val-result-label"),
                    Span(id="rev-ev", cls="val-result-value"), cls="val-result"),
                cls="val-method",
            ),
            Div(
                H4("EV / EBITDA Multiple", cls="val-method-title"),
                _slider("ebitda_multiple", "EV / EBITDA", round(ebitda_mult_default, 1), 1.0, 50.0, 0.5),
                Div(Span("Enterprise Value", cls="val-result-label"),
                    Span(id="ebitda-ev", cls="val-result-value"), cls="val-result"),
                cls="val-method",
            ),
            Div(
                H4("EV / EBIT Multiple", cls="val-method-title"),
                _slider("ebit_multiple", "EV / EBIT", round(ebit_mult_default, 1), 1.0, 60.0, 0.5),
                Div(Span("Enterprise Value", cls="val-result-label"),
                    Span(id="ebit-ev", cls="val-result-value"), cls="val-result"),
                cls="val-method",
            ),
            cls="val-row val-row-3",
        )

        dcf_section = Div(
            H4("DCF Valuation", cls="val-method-title"),
            Div(
                _slider("dcf_growth", "Revenue Growth", max(growth, 5.0), 0.0, 40.0, 0.5),
                _slider("dcf_wacc", "WACC", 12.0, 5.0, 30.0, 0.5),
                _slider("dcf_terminal", "Terminal Growth", 2.0, 0.0, 8.0, 0.5),
                _slider("dcf_years", "Projection Years", 5, 3, 10, 1),
                _slider("dcf_capex", "CapEx (% Revenue)", 3.0, 0.0, 15.0, 0.5),
                _slider("dcf_tax", "Tax Rate", 20.0, 0.0, 40.0, 1.0),
                cls="val-inputs val-inputs-grid",
            ),
            Div(Span("Enterprise Value", cls="val-result-label"),
                Span(id="dcf-ev", cls="val-result-value"), cls="val-result"),
            cls="val-method",
        )

        wacc_section = Div(
            H4("WACC Calculator", cls="val-method-title"),
            Div(
                _slider("wacc_rfr", "Risk-Free Rate", 3.5, 0.0, 10.0, 0.1),
                _slider("wacc_beta", "Levered Beta", 1.1, 0.3, 3.0, 0.05),
                _slider("wacc_mrp", "Market Risk Premium", 5.7, 3.0, 10.0, 0.1),
                _slider("wacc_country", "Country Risk Premium", 0.7, 0.0, 8.0, 0.1),
                _slider("wacc_size", "Size Premium", 5.0, 0.0, 15.0, 0.5),
                _slider("wacc_de", "D/E Ratio", 0.3, 0.0, 3.0, 0.05),
                _slider("wacc_cod", "Cost of Debt (pre-tax)", 5.0, 1.0, 15.0, 0.5),
                _slider("wacc_taxr", "Tax Rate", 20.0, 0.0, 40.0, 1.0),
                cls="val-inputs val-inputs-grid-4",
            ),
            Div(
                Div(Span("Cost of Equity", cls="val-result-label"),
                    Span(id="wacc-ke", cls="val-result-value val-result-sub"), cls="val-result"),
                Div(Span("WACC", cls="val-result-label"),
                    Span(id="wacc-result", cls="val-result-value"), cls="val-result"),
                Button("Apply to DCF", cls="val-apply-btn", onclick="applyWACC()"),
                cls="val-wacc-footer",
            ),
            cls="val-method",
        )

        bridge_section = Div(
            H4("Equity Bridge", cls="val-method-title"),
            Div(
                Div(
                    _slider("bridge_cash", f"Cash ({sym}K)", 0.0, 0.0, 10000.0, 50.0),
                    _slider("bridge_debt", f"Total Debt ({sym}K)", 0.0, 0.0, 10000.0, 50.0),
                    _slider("bridge_minority", "Minority Interest (%)", 0.0, 0.0, 30.0, 0.5),
                    cls="val-inputs val-inputs-grid",
                ),
                Div(
                    Div(Span("Avg. Enterprise Value", cls="val-result-label"),
                        Span(id="bridge-ev", cls="val-result-value val-result-sub"), cls="val-result"),
                    Div(Span("Equity Value", cls="val-result-label"),
                        Span(id="bridge-equity", cls="val-result-value"), cls="val-result"),
                    cls="val-bridge-footer",
                ),
                cls="val-method",
            ),
            cls="val-bridge-row",
        )

        chart_section = Div(
            H4("Valuation Comparison", cls="val-method-title"),
            Div(id="val-chart", cls="val-chart-container"),
            Div(Button("Download XLS", cls="val-export-btn", onclick="exportValuation()"),
                cls="val-export-row"),
            cls="val-method",
        )

        industry_section = Div(
            H4("Industry Benchmark", cls="val-method-title"),
            Select(*industry_options, id="industry_select", cls="val-industry-select",
                   onchange="updateIndustryMultiples()"),
            cls="val-method",
        )

        simulator = Div(
            industry_section, multiples_row, dcf_section, wacc_section,
            bridge_section, chart_section,
            cls="val-simulator", id="val-simulator",
        )

        rev_multiples_json = json.dumps(REVENUE_MULT)
        ebitda_multiples_json = json.dumps(EBITDA_MULT)

        sim_js = f"""
<script src="https://cdn.plot.ly/plotly-2.35.2.min.js" charset="utf-8"></script>
<script>
(function() {{
    var SYM = "{sym}";
    var REV = {rev};
    var EBITDA = {ebitda};
    var EBIT = {ebit};
    var COMPANY = {json.dumps(selected["name"])};
    var SLUG = {json.dumps(selected["slug"])};
    var REV_MULT = {rev_multiples_json};
    var EBITDA_MULT = {ebitda_multiples_json};

    function fmt(v) {{
        if (v < 0) return "-" + fmt(-v);
        if (v >= 1e9) return SYM + (v/1e9).toFixed(2) + "B";
        if (v >= 1e6) return SYM + (v/1e6).toFixed(2) + "M";
        if (v >= 1e3) return SYM + (v/1e3).toFixed(0) + "K";
        return SYM + v.toFixed(0);
    }}

    function getVal(id) {{ return parseFloat(document.getElementById(id).value); }}
    function setLabel(id, text) {{ var el = document.getElementById(id); if(el) el.textContent = text; }}

    window.updateIndustryMultiples = function() {{
        var ind = document.getElementById("industry_select").value;
        if (REV_MULT[ind]) {{
            var s = document.getElementById("rev_multiple");
            s.value = Math.min(REV_MULT[ind], parseFloat(s.max));
        }}
        if (EBITDA_MULT[ind]) {{
            var m = EBITDA_MULT[ind];
            if (m.ebitda) {{
                var s2 = document.getElementById("ebitda_multiple");
                s2.value = Math.min(m.ebitda, parseFloat(s2.max));
            }}
            if (m.ebit) {{
                var s3 = document.getElementById("ebit_multiple");
                s3.value = Math.min(m.ebit, parseFloat(s3.max));
            }}
        }}
        calc();
    }};

    window.applyWACC = function() {{
        var wacc = calcWACC();
        document.getElementById("dcf_wacc").value = wacc.toFixed(1);
        calc();
    }};

    function calcWACC() {{
        var rfr = getVal("wacc_rfr") / 100;
        var beta = getVal("wacc_beta");
        var mrp = getVal("wacc_mrp") / 100;
        var country = getVal("wacc_country") / 100;
        var size = getVal("wacc_size") / 100;
        var de = getVal("wacc_de");
        var cod = getVal("wacc_cod") / 100;
        var tax = getVal("wacc_taxr") / 100;
        var ke = rfr + beta * mrp + country + size;
        var ed = 1 / (1 + de);
        var dd = de / (1 + de);
        var wacc = ke * ed + cod * (1 - tax) * dd;
        setLabel("wacc-ke", (ke * 100).toFixed(1) + "%");
        setLabel("wacc-result", (wacc * 100).toFixed(1) + "%");

        setLabel("wacc_rfr_val", (rfr*100).toFixed(1) + "%");
        setLabel("wacc_beta_val", beta.toFixed(2));
        setLabel("wacc_mrp_val", (mrp*100).toFixed(1) + "%");
        setLabel("wacc_country_val", (country*100).toFixed(1) + "%");
        setLabel("wacc_size_val", (size*100).toFixed(1) + "%");
        setLabel("wacc_de_val", de.toFixed(2) + "x");
        setLabel("wacc_cod_val", (cod*100).toFixed(1) + "%");
        setLabel("wacc_taxr_val", (tax*100).toFixed(0) + "%");
        return wacc * 100;
    }}

    function calc() {{
        var rm = getVal("rev_multiple");
        setLabel("rev_multiple_val", rm.toFixed(1) + "x");
        var revEV = REV * rm;
        setLabel("rev-ev", fmt(revEV));

        var em = getVal("ebitda_multiple");
        setLabel("ebitda_multiple_val", em.toFixed(1) + "x");
        var ebitdaEV = EBITDA * em;
        setLabel("ebitda-ev", fmt(ebitdaEV));

        var xm = getVal("ebit_multiple");
        setLabel("ebit_multiple_val", xm.toFixed(1) + "x");
        var ebitEV = EBIT * xm;
        setLabel("ebit-ev", fmt(ebitEV));

        var g = getVal("dcf_growth") / 100;
        var wacc = getVal("dcf_wacc") / 100;
        var tg = getVal("dcf_terminal") / 100;
        var n = parseInt(document.getElementById("dcf_years").value);
        var capex = getVal("dcf_capex") / 100;
        var taxRate = getVal("dcf_tax") / 100;
        setLabel("dcf_growth_val", (g*100).toFixed(1) + "%");
        setLabel("dcf_wacc_val", (wacc*100).toFixed(1) + "%");
        setLabel("dcf_terminal_val", (tg*100).toFixed(1) + "%");
        setLabel("dcf_years_val", n + "yr");
        setLabel("dcf_capex_val", (capex*100).toFixed(1) + "%");
        setLabel("dcf_tax_val", (taxRate*100).toFixed(0) + "%");

        var baseFCF = EBITDA > 0 ? EBITDA * (1 - taxRate) - REV * capex : REV * 0.05;
        var fcf = baseFCF;
        var pv = 0;
        for (var i = 1; i <= n; i++) {{
            fcf *= (1 + g);
            pv += fcf / Math.pow(1 + wacc, i);
        }}
        var tv = (wacc > tg) ? (fcf * (1 + tg)) / (wacc - tg) : fcf * 20;
        var pvTV = tv / Math.pow(1 + wacc, n);
        var dcfEV = pv + pvTV;
        setLabel("dcf-ev", fmt(dcfEV));

        calcWACC();

        var cash = getVal("bridge_cash") * 1000;
        var debt = getVal("bridge_debt") * 1000;
        var minority = getVal("bridge_minority") / 100;
        setLabel("bridge_cash_val", SYM + (getVal("bridge_cash")).toFixed(0) + "K");
        setLabel("bridge_debt_val", SYM + (getVal("bridge_debt")).toFixed(0) + "K");
        setLabel("bridge_minority_val", (minority * 100).toFixed(1) + "%");

        var avgEV = (revEV + ebitdaEV + ebitEV + dcfEV) / 4;
        var equity = avgEV + cash - debt;
        equity *= (1 - minority);
        setLabel("bridge-ev", fmt(avgEV));
        setLabel("bridge-equity", fmt(equity));

        var methods = ["EV/Revenue", "EV/EBITDA", "EV/EBIT", "DCF"];
        var values = [revEV, ebitdaEV, ebitEV, dcfEV];
        var colors = ["#F59E0B", "#FBBF24", "#D97706", "#B45309"];

        Plotly.newPlot("val-chart", [{{
            type: "bar",
            x: methods,
            y: values,
            marker: {{ color: colors, line: {{ width: 0 }} }},
            text: values.map(function(v) {{ return fmt(v); }}),
            textposition: "outside",
            textfont: {{ family: "JetBrains Mono", size: 12, color: "#d1d5db" }},
            hovertemplate: "%{{x}}: %{{text}}<extra></extra>",
        }}], {{
            paper_bgcolor: "rgba(0,0,0,0)",
            plot_bgcolor: "rgba(0,0,0,0)",
            font: {{ color: "#d1d5db", family: "Inter" }},
            margin: {{ t: 30, b: 40, l: 50, r: 20 }},
            yaxis: {{
                gridcolor: "rgba(255,255,255,0.08)",
                tickformat: ",.0f",
                title: {{ text: "Enterprise Value (" + SYM + ")" }},
            }},
            xaxis: {{ fixedrange: true }},
            height: 300,
            showlegend: false,
            shapes: [{{
                type: "line", y0: avgEV, y1: avgEV, x0: -0.5, x1: 3.5,
                line: {{ color: "#F59E0B", width: 2, dash: "dash" }},
            }}],
            annotations: [{{
                x: 3.5, y: avgEV, text: "Avg: " + fmt(avgEV),
                showarrow: false, font: {{ color: "#F59E0B", size: 11 }},
                xanchor: "right", yanchor: "bottom",
            }}],
        }}, {{ responsive: true, displayModeBar: false }});
    }}

    window.exportValuation = function() {{
        var params = {{
            company: SLUG,
            rev_multiple: getVal("rev_multiple"),
            ebitda_multiple: getVal("ebitda_multiple"),
            ebit_multiple: getVal("ebit_multiple"),
            dcf_growth: getVal("dcf_growth"),
            dcf_wacc: getVal("dcf_wacc"),
            dcf_terminal: getVal("dcf_terminal"),
            dcf_years: parseInt(document.getElementById("dcf_years").value),
            dcf_capex: getVal("dcf_capex"),
            dcf_tax: getVal("dcf_tax"),
            wacc_rfr: getVal("wacc_rfr"),
            wacc_beta: getVal("wacc_beta"),
            wacc_mrp: getVal("wacc_mrp"),
            wacc_country: getVal("wacc_country"),
            wacc_size: getVal("wacc_size"),
            wacc_de: getVal("wacc_de"),
            wacc_cod: getVal("wacc_cod"),
            wacc_taxr: getVal("wacc_taxr"),
            bridge_cash: getVal("bridge_cash"),
            bridge_debt: getVal("bridge_debt"),
            bridge_minority: getVal("bridge_minority"),
            industry: document.getElementById("industry_select").value,
        }};
        var form = document.createElement("form");
        form.method = "POST"; form.action = "/app/valuation/export";
        form.style.display = "none";
        for (var k in params) {{
            var inp = document.createElement("input");
            inp.name = k; inp.value = params[k]; form.appendChild(inp);
        }}
        document.body.appendChild(form); form.submit();
        document.body.removeChild(form);
    }};

    document.querySelectorAll(".val-slider").forEach(function(s) {{
        s.addEventListener("input", calc);
    }});
    calc();
}})();
</script>"""

    body = Body(
        Div(id="left-overlay", cls="left-overlay", onclick="toggleLeftPane()"),
        left_pane(user_email=user_email, current_path="/app/valuation",
                  current_role=current_role),
        Div(
            Div(
                Div(
                    Button("☰", cls="mobile-menu-btn", onclick="toggleLeftPane()"),
                    Span("Valuation Simulator", cls="chat-header-title"),
                    cls="chat-header-left",
                ),
                cls="chat-header",
            ),
            Div(selector, summary, simulator, cls="companies-wrap"),
            cls="center-pane pipeline-center",
        ),
        Script(src="/chat.js"),
        NotStr(sim_js),
        Script(NotStr(_autocomplete_js())),
        cls="app pipeline-app",
    )

    head = Head(
        Meta(charset="utf-8"),
        Meta(name="viewport", content="width=device-width, initial-scale=1"),
        Title("Valuation Simulator · LiquidRound"),
        Link(rel="icon", type="image/svg+xml", href="/favicon.svg"),
        Link(rel="preconnect", href="https://fonts.googleapis.com"),
        Link(rel="preconnect", href="https://fonts.gstatic.com", crossorigin=""),
        Link(rel="stylesheet",
             href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap"),
        Script(src="https://cdn.tailwindcss.com"),
        Link(rel="stylesheet", href="/app.css"),
        Link(rel="stylesheet", href="/valuation.css"),
        Script(src="https://cdn.plot.ly/plotly-2.35.2.min.js"),
    )
    return Html(head, body)


@ar.route("/app/valuation/search")
def valuation_search(q: str = ""):
    q = q.strip()
    if len(q) < 1:
        return JSONResponse([])
    sym = "€"
    rows = _fetch_all(
        f"SELECT slug, name, sector, revenue_ltm, hq_city "
        f"FROM {DB_SCHEMA}.companies "
        f"WHERE revenue_ltm > 0 AND (name ILIKE %s OR slug ILIKE %s OR hq_city ILIKE %s) "
        f"ORDER BY revenue_ltm DESC LIMIT 12",
        (f"%{q}%", f"%{q}%", f"%{q}%"),
    )
    results = [{
        "slug": r["slug"],
        "label": f"{r['name']} — {sym}{_f(r['revenue_ltm'])/1e6:.1f}M",
        "sub": f"{r.get('hq_city') or ''} · {(r.get('sector') or '').replace('_', ' ').title()}",
    } for r in rows]
    return JSONResponse(results)


@ar.route("/app/valuation/export", methods=["POST"])
async def valuation_export(request: Request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    form = await request.form()
    p = {k: form.get(k, "") for k in form}

    slug = p.get("company", "")
    co = _fetch_one(
        f"SELECT name, sector, revenue_ltm, ebitda_ltm, ebitda_margin, "
        f"growth_rate, employees, hq_city, country "
        f"FROM {DB_SCHEMA}.companies WHERE slug = %s", (slug,)
    ) if slug else None

    rev = _f(co["revenue_ltm"]) if co else 0
    ebitda = _f(co["ebitda_ltm"]) if co else 0
    ebit = ebitda * 0.85
    name = co["name"] if co else "Company"

    rm = float(p.get("rev_multiple", 2.0))
    em = float(p.get("ebitda_multiple", 10.0))
    xm = float(p.get("ebit_multiple", 12.0))
    dcf_g = float(p.get("dcf_growth", 5.0)) / 100
    dcf_w = float(p.get("dcf_wacc", 12.0)) / 100
    dcf_tg = float(p.get("dcf_terminal", 2.0)) / 100
    dcf_n = int(float(p.get("dcf_years", 5)))
    dcf_capex = float(p.get("dcf_capex", 3.0)) / 100
    dcf_tax = float(p.get("dcf_tax", 20.0)) / 100

    wacc_rfr = float(p.get("wacc_rfr", 3.5)) / 100
    wacc_beta = float(p.get("wacc_beta", 1.1))
    wacc_mrp = float(p.get("wacc_mrp", 5.7)) / 100
    wacc_country = float(p.get("wacc_country", 0.7)) / 100
    wacc_size = float(p.get("wacc_size", 5.0)) / 100
    wacc_de = float(p.get("wacc_de", 0.3))
    wacc_cod = float(p.get("wacc_cod", 5.0)) / 100
    wacc_taxr = float(p.get("wacc_taxr", 20.0)) / 100

    cash = float(p.get("bridge_cash", 0)) * 1000
    debt = float(p.get("bridge_debt", 0)) * 1000
    minority = float(p.get("bridge_minority", 0)) / 100

    hdr_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    hdr_font = Font(bold=True, color="000000", size=11)
    num_fmt = "#,##0"
    pct_fmt = "0.0%"

    wb = Workbook()

    ws = wb.active
    ws.title = "Valuation Summary"
    ws.append(["Valuation Method", "Enterprise Value"])
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font

    rev_ev = rev * rm
    ebitda_ev = ebitda * em
    ebit_ev = ebit * xm

    base_fcf = ebitda * (1 - dcf_tax) - rev * dcf_capex if ebitda > 0 else rev * 0.05
    fcf = base_fcf
    pv = 0
    for i in range(1, dcf_n + 1):
        fcf *= (1 + dcf_g)
        pv += fcf / (1 + dcf_w) ** i
    tv = (fcf * (1 + dcf_tg)) / (dcf_w - dcf_tg) if dcf_w > dcf_tg else fcf * 20
    pv_tv = tv / (1 + dcf_w) ** dcf_n
    dcf_ev = pv + pv_tv

    avg_ev = (rev_ev + ebitda_ev + ebit_ev + dcf_ev) / 4
    equity = avg_ev + cash - debt
    equity *= (1 - minority)

    ws.append(["EV/Revenue", rev_ev])
    ws.append(["EV/EBITDA", ebitda_ev])
    ws.append(["EV/EBIT", ebit_ev])
    ws.append(["DCF", dcf_ev])
    ws.append(["Average", avg_ev])
    ws.append([])
    ws.append(["(+) Cash", cash])
    ws.append(["(-) Debt", debt])
    ws.append(["(-) Minority Interest", minority])
    ws.append(["Equity Value", equity])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            if isinstance(cell.value, (int, float)) and abs(cell.value) > 1:
                cell.number_format = num_fmt
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18

    ws2 = wb.create_sheet("Multiples")
    ws2.append(["Company", name])
    ws2.append(["Industry", p.get("industry", "")])
    ws2.append([])
    ws2.append(["Metric", "Value", "Multiple", "Enterprise Value"])
    for cell in ws2[4]:
        cell.fill = hdr_fill
        cell.font = hdr_font
    ws2.append(["Revenue (LTM)", rev, rm, rev_ev])
    ws2.append(["EBITDA (LTM)", ebitda, em, ebitda_ev])
    ws2.append(["EBIT (est.)", ebit, xm, ebit_ev])
    for row in ws2.iter_rows(min_row=5, min_col=2, max_col=4):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = num_fmt if abs(cell.value) > 10 else "0.0"
    for c in "ABCD":
        ws2.column_dimensions[c].width = 20

    ws3 = wb.create_sheet("DCF")
    ws3.append(["DCF Valuation — " + name])
    ws3["A1"].font = Font(bold=True, size=12)
    ws3.append([])
    ws3.append(["Assumptions"])
    ws3["A3"].font = Font(bold=True, size=10)
    ws3.append(["Revenue Growth", dcf_g])
    ws3.append(["WACC (Discount Rate)", dcf_w])
    ws3.append(["Terminal Growth", dcf_tg])
    ws3.append(["Projection Years", dcf_n])
    ws3.append(["CapEx (% Revenue)", dcf_capex])
    ws3.append(["Tax Rate", dcf_tax])
    for r in range(4, 10):
        ws3.cell(r, 2).number_format = pct_fmt if r != 7 else "0"

    ws3.append([])
    headers = ["Period"] + [f"Year {i}" for i in range(1, dcf_n + 1)] + ["Terminal"]
    ws3.append(headers)
    for cell in ws3[ws3.max_row]:
        cell.fill = hdr_fill
        cell.font = hdr_font

    fcf = base_fcf
    fcf_row = ["Free Cash Flow"]
    disc_row = ["Discounted FCF"]
    for i in range(1, dcf_n + 1):
        fcf *= (1 + dcf_g)
        fcf_row.append(fcf)
        disc_row.append(fcf / (1 + dcf_w) ** i)
    fcf_row.append(tv)
    disc_row.append(pv_tv)
    ws3.append(fcf_row)
    ws3.append(disc_row)
    for row in ws3.iter_rows(min_row=ws3.max_row - 1, max_row=ws3.max_row, min_col=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = num_fmt
    ws3.append([])
    ws3.append(["Sum of Discounted FCFs", pv])
    ws3.append(["PV of Terminal Value", pv_tv])
    ws3.append(["Enterprise Value", dcf_ev])
    for r in range(ws3.max_row - 2, ws3.max_row + 1):
        ws3.cell(r, 2).number_format = num_fmt
    for c in range(1, dcf_n + 4):
        ws3.column_dimensions[ws3.cell(1, c).column_letter].width = 16

    ws4 = wb.create_sheet("WACC")
    ws4.append(["WACC Calculation"])
    ws4["A1"].font = Font(bold=True, size=12)
    ws4.append([])
    ke = wacc_rfr + wacc_beta * wacc_mrp + wacc_country + wacc_size
    ed = 1 / (1 + wacc_de)
    dd = wacc_de / (1 + wacc_de)
    wacc_val = ke * ed + wacc_cod * (1 - wacc_taxr) * dd
    ws4.append(["Component", "Value", "Source"])
    for cell in ws4[3]:
        cell.fill = hdr_fill
        cell.font = hdr_font
    ws4.append(["Risk-Free Rate", wacc_rfr, "Government bond yield"])
    ws4.append(["Levered Beta", wacc_beta, "Industry beta (Damodaran)"])
    ws4.append(["Market Risk Premium", wacc_mrp, "Historical equity premium"])
    ws4.append(["Country Risk Premium", wacc_country, "Damodaran country CRP"])
    ws4.append(["Size Premium", wacc_size, "Duff & Phelps size study"])
    ws4.append([])
    ws4.append(["Cost of Equity (Ke)", ke])
    ws4.append([])
    ws4.append(["D/E Ratio", wacc_de])
    ws4.append(["E/(D+E)", ed])
    ws4.append(["D/(D+E)", dd])
    ws4.append(["Cost of Debt (pre-tax)", wacc_cod])
    ws4.append(["Tax Rate", wacc_taxr])
    ws4.append(["After-tax Cost of Debt", wacc_cod * (1 - wacc_taxr)])
    ws4.append([])
    ws4.append(["WACC", wacc_val])
    ws4[f"A{ws4.max_row}"].font = Font(bold=True, size=11)
    ws4[f"B{ws4.max_row}"].font = Font(bold=True, size=11)
    for r in range(4, ws4.max_row + 1):
        cell = ws4.cell(r, 2)
        if isinstance(cell.value, (int, float)):
            cell.number_format = pct_fmt if abs(cell.value) < 1 else "0.00"
    for c in "ABC":
        ws4.column_dimensions[c].width = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = "".join(c if c.isascii() and c.isalnum() or c in "-_ " else "" for c in name).strip().replace(" ", "-")[:30]
    filename = f"LiquidRound-Valuation-{safe_name}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
